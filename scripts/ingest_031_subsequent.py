#!/usr/bin/env python3
"""031 직후계좌(영장 회신 — 계좌 직후 이체·집금) → 그래프 적재 (결정론, LLM 무관).

030(영장계좌=1차 대포계좌)에서 자금이 빠져나간 "직후계좌"의 흐름을 담는다.
031 폴더의 구조화 데이터 2종만 결정론 파싱(나머지는 PDF/역IP 포렌식 → 별도 트랙):

  ① 스마트폰뱅킹(...).xlsx  — 하나 당행이체 내역
     출금계좌 → vt_bacnt(본계좌)  ; 입금계좌 → vt_bacnt(수취인)
     (본)-[transferred_to {out, txn_count, total_amount, first/last_dt}]->(수취인)   자금흐름 OUT
  ② 기지국 수사를 위한 자료...xlsx  — CD현금 입금 지점 피벗(입금횟수·입금총액)
     거래점 → vt_atm {atm_nm, addr, deposit_count, deposit_total, first/last_dt}
     (vt_atm)-[sourced_from]->(vt_src '기지국수사자료')   집금 ATM 지점망(현금 인출/입금 지리)

미적재(구조화 아님): 우리/농협 계좌 거래내역 PDF(원본 표는 PDF), 역IP 계좌(농협).xls(IP↔계좌 역조회),
IP.xls(사용이력), 텔레뱅킹(데이타없음). PDF는 표 추출 별도 필요.
멱등(MERGE). 실행: python3 scripts/ingest_031_subsequent.py --graph ep5_graph [--dry-run]
"""
import argparse
import glob
import os
import re
import sys
import unicodedata
from collections import defaultdict

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from app import create_app
from app.database import safe_set_graph_path
import psycopg2

GRAPH = 'ep5_graph'
SRC_ID = 'EP5-031'
BASE = ('/Users/iankwon/Downloads/00_종합시나리오 및 데이터셋/데이터셋')


def esc(v):
    return str(v).replace("\\", "\\\\").replace("'", "''")


def norm_num(v):
    return re.sub(r'[-\s\xa0]', '', str(v)) if v else ''


def to_int(v):
    n = re.sub(r'[^\d]', '', str(v)) if v is not None else ''
    return int(n) if n.isdigit() else 0


def find_031_dir():
    d5 = [p for p in glob.glob(BASE + '/EP5*') if os.path.isdir(p)][0]
    return [p for p in glob.glob(d5 + '/*')
            if '031' in unicodedata.normalize('NFC', os.path.basename(p))][0]


def find_files(d31, needle):
    out = []
    for root, _, files in os.walk(d31):
        for f in files:
            fn = unicodedata.normalize('NFC', f)
            if fn.startswith(('._', '~')):
                continue
            if needle in fn and fn.endswith(('.xls', '.xlsx')):
                out.append(os.path.join(root, f))
    return out


def parse_smartbank(path, nodes, edges):
    """스마트폰뱅킹 이체내역 → 본계좌→수취인 transferred_to(out) 집계."""
    import openpyxl
    ws = openpyxl.load_workbook(path, data_only=True).active
    rows = list(ws.iter_rows(values_only=True))
    h = [str(c).strip() if c else '' for c in rows[0]]
    def col(*kw):
        for i, x in enumerate(h):
            if any(k in x for k in kw):
                return i
        return None
    c_out = col('출금계좌'); c_amt = col('출금금액'); c_in = col('입금계좌')
    c_rcv = col('수취인'); c_dt = col('이체일자')
    if c_out is None or c_in is None:
        return 0
    agg = defaultdict(lambda: {'n': 0, 'amt': 0, 'first': '', 'last': '', 'nm': '', 'own': ''})
    for r in rows[1:]:
        def g(i):
            return r[i] if i is not None and i < len(r) else None
        own = norm_num(g(c_out)); peer = norm_num(g(c_in))
        if not own or not peer or len(peer) < 6:
            continue
        amt = to_int(g(c_amt)); nm = str(g(c_rcv)).strip() if g(c_rcv) else ''
        dt = str(g(c_dt)).strip()[:10] if g(c_dt) else ''
        a = agg[(own, peer)]
        a['n'] += 1; a['amt'] += amt; a['nm'] = nm or a['nm']; a['own'] = own
        a['first'] = min(a['first'] or dt, dt) if dt else a['first']
        a['last'] = max(a['last'], dt) if dt else a['last']
    n = 0
    for (own, peer), a in agg.items():
        ok = ('vt_bacnt', own); pk = ('vt_bacnt', peer)
        nodes.setdefault(ok, {}).update({'account_no': own, 'bank_nm': '하나', 'source_id': SRC_ID})
        nodes.setdefault(pk, {}).update({'account_no': peer, 'dpstr': a['nm'], 'source_id': SRC_ID})
        edges.append(('transferred_to', ok, pk,
                      {'txn_count': a['n'], 'total_amount': a['amt'],
                       'first_dlng_dt': a['first'], 'last_dlng_dt': a['last'],
                       'channel': '스마트폰뱅킹', 'source_id': SRC_ID}))
        # 수취인 명의 → 인물 노드 + has_account
        if a['nm'] and re.fullmatch(r'[가-힣]{2,4}', a['nm'].strip()):
            pp = ('vt_psn', a['nm'].strip())
            nodes.setdefault(pp, {'name': a['nm'].strip(), 'source_id': SRC_ID})
            edges.append(('has_account', pp, pk, {'source_id': SRC_ID}))
        n += 1
    return n


def parse_cell_pivot(path, nodes, edges):
    """기지국 자료 — 입금 피벗 시트(거래점별 입금횟수/총액) → vt_atm + sourced_from(vt_src)."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    src_k = ('vt_src', '기지국수사자료')
    nodes.setdefault(src_k, {'src_name': '기지국수사자료', 'src_type': '영장회신', 'source_id': SRC_ID})
    n = 0
    for sn in wb.sheetnames:
        if '피벗' not in sn:      # 원시 입금 891행은 제외(고카디널리티) — 피벗 요약만
            continue
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        hi = next((i for i, r in enumerate(rows[:5])
                   if r and any('거래점' in str(c) for c in r if c)), None)
        if hi is None:
            continue
        h = [str(c).strip() if c else '' for c in rows[hi]]
        def col(*kw):
            for i, x in enumerate(h):
                if any(k in x for k in kw):
                    return i
            return None
        c_br = col('거래점'); c_ad = col('주소'); c_cnt = col('입금횟수'); c_tot = col('입금총액')
        c_s = col('거래시작'); c_e = col('거래종료')
        if c_br is None:
            continue
        for r in rows[hi + 1:]:
            def g(i):
                return r[i] if i is not None and i < len(r) else None
            br = str(g(c_br)).strip() if g(c_br) else ''
            if not br or br in ('None', '합계', '총합계'):
                continue
            atm_k = ('vt_atm', br)
            nd = nodes.setdefault(atm_k, {'atm_nm': br, 'source_id': SRC_ID})
            if g(c_ad):
                nd['addr'] = str(g(c_ad)).strip()
            if g(c_cnt) is not None:
                nd['deposit_count'] = nd.get('deposit_count', 0) + to_int(g(c_cnt))
            if g(c_tot) is not None:
                nd['deposit_total'] = nd.get('deposit_total', 0) + to_int(g(c_tot))
            if g(c_s):
                nd['first_dt'] = str(g(c_s)).strip()[:12]
            if g(c_e):
                nd['last_dt'] = str(g(c_e)).strip()[:12]
            edges.append(('sourced_from', atm_k, src_k, {'source_id': SRC_ID}))
            n += 1
    return n


def build():
    d31 = find_031_dir()
    nodes, edges = {}, []
    n_sb, n_atm = 0, 0
    for p in find_files(d31, '스마트폰뱅킹'):
        n_sb += parse_smartbank(p, nodes, edges)
    for p in find_files(d31, '기지국'):
        n_atm += parse_cell_pivot(p, nodes, edges)
    return nodes, edges, n_sb, n_atm


KP = {'vt_bacnt': 'account_no', 'vt_atm': 'atm_nm', 'vt_src': 'src_name', 'vt_psn': 'name'}


def merge_cypher(nodes, edges):
    stmts = []
    for (label, key), props in nodes.items():
        setp = ', '.join(f"n.{a} = '{esc(b)}'" for a, b in props.items() if b != '')
        stmts.append(f"MERGE (n:{label} {{{KP[label]}:'{esc(props[KP[label]])}'}})" + (f" SET {setp}" if setp else ""))
    for rel, (fl, fk), (tl, tk), props in edges:
        a = f"(a:{fl} {{{KP[fl]}:'{esc(nodes[(fl,fk)][KP[fl]])}'}})"
        b = f"(b:{tl} {{{KP[tl]}:'{esc(nodes[(tl,tk)][KP[tl]])}'}})"
        sp = ', '.join(f"e.{k2} = '{esc(v2)}'" for k2, v2 in props.items() if v2 != '')
        stmts.append(f"MATCH {a}, {b} MERGE (a)-[e:{rel}]->(b)" + (f" SET {sp}" if sp else ""))
    return stmts


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--graph', default=GRAPH)
    ap.add_argument('--dry-run', action='store_true')
    args = ap.parse_args()
    nodes, edges, n_sb, n_atm = build()
    from collections import Counter
    print(f"[빌드] 스마트폰뱅킹 이체 {n_sb} · 기지국 ATM지점 {n_atm} → 노드 {len(nodes)} / 엣지 {len(edges)}")
    print("  노드:", dict(Counter(l for (l, _) in nodes)))
    print("  엣지:", dict(Counter(r for (r, *_) in edges)))
    stmts = merge_cypher(nodes, edges)
    if args.dry_run:
        for s in stmts[:6]:
            print("  ", s[:160])
        return
    app = create_app()
    with app.app_context():
        conn = psycopg2.connect(**app.config['DB_CONFIG']); conn.autocommit = False
        cur = conn.cursor()
        try:
            safe_set_graph_path(cur, args.graph)
            for vl in sorted({l for (l, _) in nodes}):
                cur.execute(f"CREATE VLABEL IF NOT EXISTS {vl};")
            for el in sorted({r for (r, *_) in edges}):
                cur.execute(f"CREATE ELABEL IF NOT EXISTS {el};")
            for s in stmts:
                cur.execute(s)
            conn.commit()
            print(f"[적재 완료] {len(stmts)}개 MERGE → {args.graph}")
        except Exception as e:
            conn.rollback(); print(f"[롤백] {e}"); raise
        finally:
            conn.close()


if __name__ == '__main__':
    main()
