#!/usr/bin/env python3
"""EP7 043 금융 연결계좌(표준화 xlsx) → 자금흐름 그래프 (결정론·LLM무관).

IP역조회로 특정된 대상자(이미나·송주혁·조승연 등)의 은행별 거래내역. 은행마다 헤더 상이
→ 헤더 자동매핑. 계좌주는 '계좌주/예금주' 컬럼 또는 상단 "예금주 : OOO" 행에서 추출.
온톨로지 매핑(V4.7):
  계좌주            → vt_psn ; (계좌주)-[has_account]->(본계좌)
  본계좌/상대계좌   → vt_bacnt {account_no, dpstr}
  입금/출금        → (방향)transferred_to {txn_count,total_amount,first/last_dt}
상대명(적요/거래내용)은 소비처·카드·이체메모가 많아 사람이름만 dpstr(030 파서와 동일 정제).
멱등(MERGE). 실행: python3 scripts/ingest_043_standard.py --graph ep7_graph [--dry-run]
"""
import argparse
import os
import re
import sys
import unicodedata
from collections import defaultdict

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from app import create_app
from app.database import safe_set_graph_path
import psycopg2

GRAPH = 'ep7_graph'
SRC = 'EP7-043'
BASE = '/Users/iankwon/Downloads/00_종합시나리오 및 데이터셋/데이터셋'
BANKS = ['국민은행 표준화', '기업은행 거래내역(표준', '신한은행 거래내역', '2차 계좌 표준화']
BIZ = re.compile(r'(카드|체크|마트|편의점|GS25|CU|세븐|떡볶이|치킨|커피|카페|Toss|토스|페이|웹|주유|약국|병원|미용|네일|헤어|보험|화재|생명|한전|가스|통신|월세|렌탈|현금|ATM|CD|이체|입금|출금|급여|이자|수수료|대체|송금|타행|당행|정기|저축|상환|납부|환급|정산|자동|펀드|대출|모바일|인터넷|폰뱅|텔레뱅|스마트폰|무통장|공과금)')
BANKPFX = re.compile(r'^(기업|국민|신한|농협|우리|하나|신협|새마을|씨티|SC|카카오|케이|토스|부산|경남|대구|광주|전북|제주|수협|산업)')


def N(p):
    return unicodedata.normalize('NFC', str(p))


def esc(v):
    return str(v).replace("\\", "\\\\").replace("'", "''")


def nn(v):
    return re.sub(r'[^\d]', '', str(v)) if v else ''


def to_int(v):
    n = nn(v)
    return int(n) if n.isdigit() else 0


def clean_peer(nm):
    j = (nm or '').strip()
    if not j or BIZ.search(j):
        return None
    core = BANKPFX.sub('', j).strip()
    return core if re.fullmatch(r'[가-힣]{2,4}', core) else None


def find_files(d7):
    picked = {}
    for r, _, fs in os.walk(d7):
        if '043' not in N(r):
            continue
        for f in fs:
            fn = N(f)
            if not fn.endswith('.xlsx') or f.startswith(('.', '~')):
                continue
            for b in BANKS:
                if b in fn:
                    prio = 1 if '_v3' in fn else 0
                    if b not in picked or prio > picked[b][0]:
                        picked[b] = (prio, os.path.join(r, f))
    return [v[1] for v in picked.values()]


def map_hdr(rows):
    hi = max(range(min(6, len(rows))), key=lambda i: sum(
        1 for c in rows[i] if c and any(k in str(c) for k in
        ['계좌주', '예금주', '거래날짜', '거래일자', '금액', '상대계좌', '구분', '적요', '거래내용'])))
    h = [str(c).strip() if c else '' for c in rows[hi]]

    def f(*kw):
        for i, x in enumerate(h):
            if any(k in x for k in kw):
                return i
        return None
    cm = {
        'acct': f('계좌') if f('계좌') is not None and '상대' not in h[f('계좌')] else None,
        'owner': f('계좌주', '예금주'),
        'date': f('거래날짜', '거래일자'),
        'gubun': f('구분', '입지'),
        'amt': f('거래금액', '금액'),
        'peer_acct': f('상대계좌'),
        'peer_nm': f('적요', '거래내용'),
    }
    # 기업형: 헤더 첫 셀이 본계좌번호(숫자12+)이고 구분 헤더 없음 → A열 데이터가 출금/입금 구분
    if cm['gubun'] is None and h and re.fullmatch(r'\d{10,}', re.sub(r'\D', '', h[0])):
        cm['gubun'] = 0
    return cm, hi


def owner_from_top(rows, hi):
    """상단 예금주행에서 (명의, 본계좌번호) 추출. 신한형 '예금주 : 조승연 / 계좌번호 : 110…'."""
    owner, acct = None, None
    for r in rows[:hi]:
        cells = [str(c) if c is not None else '' for c in (r or [])]
        joined = ' '.join(cells)
        if owner is None:
            m = re.search(r'예금주\s*[:：]?\s*([가-힣]{2,4})', joined)
            if m:
                owner = m.group(1)
            else:  # '예금주 :'(A셀) + '조승연'(B셀) 별도 셀
                for j, c in enumerate(cells):
                    if '예금주' in c and j + 1 < len(cells) and re.fullmatch(r'[가-힣]{2,4}', cells[j + 1].strip()):
                        owner = cells[j + 1].strip(); break
        if acct is None:  # '계좌번호 :' 옆의 본계좌번호
            am = re.search(r'계좌번호\s*[:：]?\s*(\d[\d\-]{7,})', joined)
            if am:
                acct = re.sub(r'\D', '', am.group(1))
            else:
                for j, c in enumerate(cells):
                    if '계좌번호' in c and j + 1 < len(cells):
                        a2 = re.sub(r'\D', '', cells[j + 1])
                        if len(a2) >= 8:
                            acct = a2; break
    return owner, acct


def build(paths):
    import openpyxl
    nodes, edges = {}, []

    def node(l, k, pr):
        nk = (l, k); nodes.setdefault(nk, {})
        nodes[nk].update({a: b for a, b in pr.items() if b != ''}); return nk

    for p in paths:
        wb = openpyxl.load_workbook(p, data_only=True)
        for sn in wb.sheetnames:
            ws = wb[sn]
            if ws.sheet_state != 'visible':
                continue
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 3:
                continue
            cm, hi = map_hdr(rows)
            if cm['peer_acct'] is None or cm['amt'] is None:
                continue
            top_owner, top_acct = owner_from_top(rows, hi)
            agg = defaultdict(lambda: {'n': 0, 'amt': 0, 'first': '', 'last': '', 'pnm': '', 'own': '', 'oacct': ''})
            for r in rows[hi + 1:]:
                def cell(k):
                    i = cm.get(k)
                    return r[i] if i is not None and i < len(r) else None
                pa = nn(cell('peer_acct'))
                if not pa or not pa.isdigit() or len(pa) < 8:
                    continue
                owner = (str(cell('owner')).strip() if cm['owner'] is not None and cell('owner') else '') or top_owner or ''
                oacct = nn(cell('acct')) if cm['acct'] is not None else (top_acct or '')
                amt = to_int(cell('amt'))
                g = str(cell('gubun') or '')
                direction = 'in' if ('입' in g) else ('out' if ('출' in g or '지급' in g) else '')
                if not direction or amt <= 0:
                    continue
                dt = str(cell('date')).strip()[:10] if cell('date') else ''
                pnm = clean_peer(str(cell('peer_nm') or ''))
                a = agg[(owner, oacct, pa, direction)]
                a['n'] += 1; a['amt'] += amt; a['pnm'] = pnm or a['pnm']
                a['first'] = min(a['first'] or dt, dt) if dt else a['first']
                a['last'] = max(a['last'], dt) if dt else a['last']
            for (owner, oacct, pa, direction), a in agg.items():
                if not owner:
                    continue
                okey = oacct or f'명의:{owner}'
                ok = node('vt_bacnt', okey, {'account_no': okey, 'dpstr': owner, 'source_id': SRC})
                own_psn = node('vt_psn', owner, {'name': owner, 'source_id': SRC})
                edges.append(('has_account', own_psn, ok, {'source_id': SRC}))
                pk = node('vt_bacnt', pa, {'account_no': pa, 'dpstr': a['pnm'], 'source_id': SRC})
                if a['pnm']:
                    pp = node('vt_psn', a['pnm'], {'name': a['pnm'], 'source_id': SRC})
                    edges.append(('has_account', pp, pk, {'source_id': SRC}))
                props = {'txn_count': a['n'], 'total_amount': a['amt'],
                         'first_dlng_dt': a['first'], 'last_dlng_dt': a['last'], 'source_id': SRC}
                if direction == 'in':
                    edges.append(('transferred_to', pk, ok, props))
                else:
                    edges.append(('transferred_to', ok, pk, props))
    return nodes, edges


KP = {'vt_bacnt': 'account_no', 'vt_psn': 'name'}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--graph', default=GRAPH)
    ap.add_argument('--dry-run', action='store_true')
    args = ap.parse_args()
    d7 = [os.path.join(BASE, x) for x in os.listdir(BASE) if N(x).startswith('EP7.')][0]
    paths = find_files(d7)
    print('[대상]', [N(os.path.basename(p)) for p in paths])
    nodes, edges = build(paths)
    from collections import Counter
    print(f"[빌드] 노드 {len(nodes)} / 엣지 {len(edges)}")
    print("  노드:", dict(Counter(l for (l, _) in nodes)))
    print("  엣지:", dict(Counter(r for (r, *_) in edges)))
    stmts = []
    for (l, k), pr in nodes.items():
        sp = ', '.join(f"n.{a} = '{esc(b)}'" for a, b in pr.items() if b != '')
        stmts.append(f"MERGE (n:{l} {{{KP[l]}:'{esc(pr[KP[l]])}'}})" + (f" SET {sp}" if sp else ""))
    for rel, (fl, fk), (tl, tk), pr in edges:
        a = f"(a:{fl} {{{KP[fl]}:'{esc(nodes[(fl,fk)][KP[fl]])}'}})"
        b = f"(b:{tl} {{{KP[tl]}:'{esc(nodes[(tl,tk)][KP[tl]])}'}})"
        sp = ', '.join(f"e.{k2} = '{esc(v2)}'" for k2, v2 in pr.items() if v2 != '')
        stmts.append(f"MATCH {a}, {b} MERGE (a)-[e:{rel}]->(b)" + (f" SET {sp}" if sp else ""))
    if args.dry_run:
        owners = sorted({p['name'] for (l, _), p in nodes.items() if l == 'vt_psn'})
        print("  계좌주/인물:", owners[:20])
        return
    app = create_app()
    with app.app_context():
        conn = psycopg2.connect(**app.config['DB_CONFIG']); conn.autocommit = False
        cur = conn.cursor()
        try:
            safe_set_graph_path(cur, args.graph)
            for vl in sorted({l for (l, _) in nodes}):
                cur.execute(f"CREATE VLABEL IF NOT EXISTS {vl};")
            for el in sorted({r for (r, *_) in edges}):
                cur.execute(f"CREATE ELABEL IF NOT EXISTS {el};")
            for s in stmts:
                cur.execute(s)
            conn.commit()
            print(f"[적재 완료] {len(stmts)}개 MERGE → {args.graph}")
        except Exception as e:
            conn.rollback(); print(f"[롤백] {e}"); raise
        finally:
            conn.close()


if __name__ == '__main__':
    main()
