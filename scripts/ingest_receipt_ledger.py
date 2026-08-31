#!/usr/bin/env python3
"""트랙A — 접수내역/범죄일람표(정형 xlsx) → AgensGraph 온톨로지 적재 (결정론, LLM 무관).

2차년도 EP1 접수내역의 예금주(대포통장 명의자)·계좌·피해자·수사단서를 tccop_graph_v6 에 MERGE 적재.
그래프에 IP·통화만 있고 접수내역(g1)이 미적재라 예금주(이진아 등)가 누락된 문제 해결.

온톨로지 매핑 (V4.7):
  예금주            → vt_psn {name}          -[has_account]->  vt_bacnt (명의)
  계좌              → vt_bacnt {account_no, bank_nm, dpstr}
  피해자            → vt_psn {name}          -[victim_in]->    vt_case
  신고 1건          → vt_case {flnm, occrn_dt, damage_amt, crime_site}
  사건 → 계좌/전화/ID  -[eg_used_account / eg_used_phone / eg_used_id]->  (수사단서)
  전화              → vt_telno {telno}
  카톡/네이버 계정   → vt_id {platform, id_val}

멱등: 전부 MERGE (재실행 안전). 계좌 정규화(하이픈 제거)로 기존 노드와 자동 병합.
실행: python3 scripts/ingest_receipt_ledger.py [--dry-run]
"""
import argparse
import re
import sys
import os

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from app import create_app
from app.database import safe_set_graph_path
import psycopg2

XLSX = ('/Users/iankwon/Downloads/00_종합시나리오 및 데이터셋/데이터셋/'
        'EP1. DS-01 (비식별화)/004_0000-03-29 접수내역★_비식별화.xlsx')
SHEET = '01_신고사건 15건 정리'      # 이진아 예금주 정본(단일 시트, 내부 정합 보장)
GRAPH = 'tccop_graph_v6'
SRC_ID = 'Y2-EP1-접수내역-004'       # provenance
CASE_PREFIX = 'Y2-EP1'


def esc(v):
    return str(v).replace("\\", "\\\\").replace("'", "''")


def norm_num(v):
    return re.sub(r'[-\s]', '', str(v)) if v else ''


def split_multi(v):
    if not v:
        return []
    return [x.strip() for x in re.split(r'[\n/,]+', str(v)) if x.strip()]


def id_type(v):
    """카톡/네이버 값 분류: 괄호로 감싼 값은 닉네임, 그 외는 계정 ID.
    닉네임은 식별력이 낮고(여러 계정 공유·변경 잦음) 계정 ID와 성격이 달라 구분 저장."""
    return 'nickname' if str(v).strip().startswith('(') else 'account_id'


_COL_KEYS = {   # 논리 컬럼 → 헤더 키워드
    'seq': ['연번'], 'occrn': ['범행일'], 'site': ['사이트'], 'victim': ['피해자'],
    'dpstr': ['예금주'], 'bank': ['은행'], 'acct': ['계좌'], 'amt': ['편취'],
    'phone': ['전화'], 'kakao': ['카카오', '카톡'], 'naver': ['네이버'],
}


def _map_columns(rows):
    """헤더 행(=헤더 키워드 셀이 가장 많은 행)을 찾아 컬럼 매핑. 병합 헤더는 위 행(대분류)으로 보완.
    데이터 행이 헤더를 덮어쓰는 문제를 방지(006 단일헤더 / 004 병합헤더 모두 대응)."""
    ncol = max(len(r) for r in rows[:4])
    _allkws = [k for kws in _COL_KEYS.values() for k in kws]
    def match_count(r):
        return sum(1 for c in r if c and any(k in str(c).replace('\n', '') for k in _allkws))
    # 동점이면 아래 행(세부 헤더) 선호 — 병합 헤더에서 대분류(위)가 잡히는 것 방지
    hdr_i = max(range(min(4, len(rows))), key=lambda i: (match_count(rows[i]), i))
    htext = [''] * ncol
    for src_i in (hdr_i - 1, hdr_i):         # 대분류(위) → 세부(주헤더)가 덮음 = 세부 우선
        if src_i < 0:
            continue
        for i, c in enumerate(rows[src_i]):
            if c and str(c).strip():
                htext[i] = str(c).strip().replace('\n', '')
    colmap = {}
    for logical, kws in _COL_KEYS.items():
        for i, h in enumerate(htext):
            if any(k in h for k in kws):
                colmap[logical] = i
                break
    sidx = colmap.get('seq', 0)
    start = hdr_i + 1
    for j in range(hdr_i + 1, len(rows)):
        v = rows[j][sidx] if sidx < len(rows[j]) else None
        if v is not None and str(v).strip().isdigit():
            start = j
            break
    return colmap, start


def build(xlsx=None, sheet=None, src_id=None, case_prefix=None):
    xlsx = xlsx or XLSX; sheet = sheet or SHEET
    src_id = src_id or SRC_ID; case_prefix = case_prefix or CASE_PREFIX
    import openpyxl
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb[sheet]
    # 숨김 시트는 작업 중간본·폐기본일 수 있어 기본 거부(사용자에게 보이는 visible 시트만 적재)
    if ws.sheet_state != 'visible' and not globals().get('_ALLOW_HIDDEN'):
        raise SystemExit(f"'{sheet}'는 숨김 시트(state={ws.sheet_state}). "
                         f"의도한 것이면 --allow-hidden. 보이는 시트: "
                         f"{[s for s in wb.sheetnames if wb[s].sheet_state=='visible']}")
    rows = list(ws.iter_rows(values_only=True))
    cm, start = _map_columns(rows)
    missing = [k for k in ('victim', 'dpstr', 'acct') if k not in cm]
    if missing:
        raise SystemExit(f"필수 컬럼 미발견: {missing} (헤더 확인 필요)")
    def cell(r, key):
        i = cm.get(key)
        return r[i] if i is not None and i < len(r) else None
    nodes, edges = {}, []   # nodes[(label,key)] = props ; edges=[(rel,(fl,fk),(tl,tk),props)]

    def node(label, key, props=None):
        k = (label, key)
        nodes.setdefault(k, {})
        if props:
            nodes[k].update({kk: vv for kk, vv in props.items() if vv})
        return k

    def edge(rel, a, b, props=None):
        edges.append((rel, a, b, props or {}))

    for r in rows[start:]:
        seq_v = cell(r, 'seq')
        if seq_v is None or not str(seq_v).strip().isdigit():
            continue
        seq = str(seq_v).strip()
        occrn = str(cell(r, 'occrn')).strip() if cell(r, 'occrn') else ''
        site = str(cell(r, 'site')).strip().replace('\n', ' ') if cell(r, 'site') else ''
        victim = str(cell(r, 'victim')).strip() if cell(r, 'victim') else ''
        dpstr = str(cell(r, 'dpstr')).strip() if cell(r, 'dpstr') else ''
        bank = str(cell(r, 'bank')).strip() if cell(r, 'bank') else ''
        acct = norm_num(cell(r, 'acct'))
        amt = norm_num(cell(r, 'amt'))
        phones = [norm_num(p) for p in split_multi(cell(r, 'phone'))]
        kakaos = split_multi(cell(r, 'kakao'))
        navers = split_multi(cell(r, 'naver'))

        flnm = f'{case_prefix}-{seq.zfill(2)}'
        case_k = node('vt_case', flnm,
                      {'flnm': flnm, 'occrn_dt': occrn,
                       'damage_amt': amt, 'crime_site': site, 'source_id': src_id})
        # 예금주(명의자) + 계좌 — 계좌 유효성: 숫자 10자리+ (범행일시 등 오탐 방지)
        if acct and acct.isdigit() and len(acct) >= 10:
            acct_k = node('vt_bacnt', acct,
                          {'account_no': acct, 'bank_nm': bank, 'dpstr': dpstr, 'source_id': src_id})
            if dpstr:
                dp_k = node('vt_psn', dpstr, {'name': dpstr, 'source_id': src_id})
                edge('has_account', dp_k, acct_k, {'source_id': src_id})
            edge('eg_used_account', case_k, acct_k, {'source_id': src_id})
        if victim:
            v_k = node('vt_psn', victim, {'name': victim, 'source_id': src_id})
            edge('victim_in', v_k, case_k, {'source_id': src_id})
        for ph in phones:
            if ph and re.fullmatch(r'0(10|70)\d{7,8}', ph):
                edge('eg_used_phone', case_k, node('vt_telno', ph, {'telno': ph, 'source_id': src_id}), {'source_id': src_id})
        for kk in kakaos:
            if kk:
                edge('eg_used_id', case_k, node('vt_id', f'kakao:{kk}',
                     {'platform': 'kakao', 'id_val': kk, 'id_type': id_type(kk), 'source_id': src_id}), {'source_id': src_id})
        for nv in navers:
            if nv:
                edge('eg_used_id', case_k, node('vt_id', f'naver:{nv}',
                     {'platform': 'naver', 'id_val': nv, 'id_type': id_type(nv), 'source_id': src_id}), {'source_id': src_id})
    return nodes, edges


_KEYPROP = {'vt_case': 'flnm', 'vt_bacnt': 'account_no', 'vt_psn': 'name',
            'vt_telno': 'telno', 'vt_id': None}   # vt_id 는 복합키(platform,id_val)


def merge_cypher(nodes, edges):
    stmts = []
    for (label, key), props in nodes.items():
        setp = ', '.join(f"n.{k} = '{esc(v)}'" for k, v in props.items())
        if label == 'vt_id':
            plat, idv = props.get('platform', ''), props.get('id_val', '')
            match = f"(n:vt_id {{platform:'{esc(plat)}', id_val:'{esc(idv)}'}})"
        else:
            kp = _KEYPROP[label]
            match = f"(n:{label} {{{kp}:'{esc(props[kp])}'}})"
        stmts.append(f"MERGE {match}" + (f" SET {setp}" if setp else ""))
    for rel, (fl, fk), (tl, tk), props in edges:
        def sel(label, props_key, node_key):
            p = nodes[(label, node_key)]
            if label == 'vt_id':
                return f"(:vt_id {{platform:'{esc(p['platform'])}', id_val:'{esc(p['id_val'])}'}})"
            kp = _KEYPROP[label]
            return f"(:{label} {{{kp}:'{esc(p[kp])}'}})"
        a = sel(fl, None, fk).replace('(:', '(a:', 1)
        b = sel(tl, None, tk).replace('(:', '(b:', 1)
        sp = ', '.join(f"e.{k} = '{esc(v)}'" for k, v in props.items())
        stmts.append(f"MATCH {a}, {b} MERGE (a)-[e:{rel}]->(b)" + (f" SET {sp}" if sp else ""))
    return stmts


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--dry-run', action='store_true')
    ap.add_argument('--graph', default=GRAPH)
    ap.add_argument('--xlsx', default=XLSX)
    ap.add_argument('--sheet', default=SHEET)
    ap.add_argument('--src-id', default=SRC_ID)
    ap.add_argument('--case-prefix', default=CASE_PREFIX)
    ap.add_argument('--allow-hidden', action='store_true', help='숨김 시트도 적재 허용')
    args = ap.parse_args()
    if args.allow_hidden:
        globals()['_ALLOW_HIDDEN'] = True

    nodes, edges = build(args.xlsx, args.sheet, args.src_id, args.case_prefix)
    from collections import Counter
    nc = Counter(l for (l, _) in nodes)
    ec = Counter(r for (r, *_ ) in edges)
    print(f"[빌드] 노드 {len(nodes)} / 엣지 {len(edges)}")
    print("  노드:", dict(nc))
    print("  엣지:", dict(ec))

    stmts = merge_cypher(nodes, edges)
    if args.dry_run:
        print(f"\n[dry-run] MERGE 문 {len(stmts)}개. 샘플 3:")
        for s in stmts[:3]:
            print("  ", s[:160])
        return

    app = create_app()
    with app.app_context():
        conn = psycopg2.connect(**app.config['DB_CONFIG'])
        conn.autocommit = False
        cur = conn.cursor()
        try:
            safe_set_graph_path(cur, args.graph)
            # AgensGraph: 사용 라벨을 먼저 선언(없으면 MERGE가 "label does not exist" 실패)
            for vl in sorted({l for (l, _) in nodes}):
                cur.execute(f"CREATE VLABEL IF NOT EXISTS {vl};")
            for el in sorted({r for (r, *_) in edges}):
                cur.execute(f"CREATE ELABEL IF NOT EXISTS {el};")
            for s in stmts:
                cur.execute(s)
            conn.commit()
            print(f"\n[적재 완료] {len(stmts)}개 MERGE → {args.graph} (트랜잭션 커밋)")
        except Exception as e:
            conn.rollback()
            print(f"\n[롤백] 오류: {e}")
            raise
        finally:
            conn.close()


if __name__ == '__main__':
    main()
