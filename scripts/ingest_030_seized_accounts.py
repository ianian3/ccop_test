#!/usr/bin/env python3
"""030 영장 계좌(압수 대상 대포통장) 거래내역 → 자금흐름 그래프 (보정판, 결정론·LLM무관).

기존 ingest_account_txn 의 EP5-030 적재 오류 교정:
  ① 본계좌 명의: 인자(--owner-name) 대신 **시트 1행 "OOO 명의 XX은행 계좌"에서 자동 추출**
     (우리은행-영장 목록 파일처럼 한 파일에 9개 계좌 시트가 있어도 시트별 명의 정확)
  ② 상대명(dpstr): '적요(의뢰인등)'에서 뽑되 **명의 오추출 차단** —
     · 사람이름(2~4자 한글)만 dpstr
     · 조직(피어스미디어·유니크프로젝트 등)은 belongs_to vt_org
     · 보험료(METLIFE·메리츠·삼성화)·공과금(한전)·코드(#·000000)·메모(３월)는 **명의 공란**
매핑(V4.7): 본/상대계좌 vt_bacnt · (방향)transferred_to · 조직 belongs_to.
대상(자동탐색): 우리은행-영장 목록(9시트)·기업 신민우·우리 문범수·국민 특이사항(다명의).
제외: 하나 스마트폰뱅킹형(031파서)·en 역조회(IP)·새마을(상대계좌번호 없음→피어스 belongs_to만).
멱등(MERGE). 실행: python3 scripts/ingest_030_seized_accounts.py --graph ep5_graph [--dry-run]
"""
import argparse
import glob
import os
import re
import sys
import unicodedata
from collections import defaultdict

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from app import create_app
from app.database import safe_set_graph_path
import psycopg2

GRAPH = 'ep5_graph'
SRC = 'EP5-030'
BASE = '/Users/iankwon/Downloads/00_종합시나리오 및 데이터셋/데이터셋'
WANT = ['우리은행-영장 목록', '신민우', '문범수', '특이사항 정리']   # 상대계좌번호 있는 형식만
INS = re.compile(r'(METLIFE|메트라이프|메리츠|삼성화|삼성생명|한화|현대해상|DB손|KB손|흥국|동양생명|라이나|미래에셋|보험|화재|생명|손해|공제|연금)')
UTIL = re.compile(r'(한전|전기|도시가스|가스|수도|상수도|통신|텔레콤|요금|관리비|국세|지방세|세금)')
ORG = re.compile(r'(미디어|프로젝트|컴퍼니|㈜|주식회사|인터내셔널|엔터|글로벌|테크|스튜디오|커머스|트레이드|캐피탈)')
BIZ = re.compile(r'(베베|헤어|월세|미용|네일|카페|마트|스토어|피부|성형|약국|의원|병원|학원|부동산|현금입금|재테크|렌탈|렌트|페이|포인트)')
BANKPFX = re.compile(r'^(기업|국민|신한|농협|우리|하나|신협|새마을|씨티|SC|카카오|케이|토스|부산|경남|대구|광주|전북|제주|수협|산업|기은|우체국)')


def N(p):
    return unicodedata.normalize('NFC', str(p))


def esc(v):
    return str(v).replace("\\", "\\\\").replace("'", "''")


def norm_num(v):
    return re.sub(r'[-\s\xa0=]', '', str(v)) if v else ''


def to_int(v):
    n = re.sub(r'[^\d]', '', str(v)) if v is not None else ''
    return int(n) if n.isdigit() else 0


def clean_peer(jeokyo):
    """적요(의뢰인등) → (사람명 or None, 조직명 or None). 노이즈는 (None,None)."""
    j = (jeokyo or '').strip()
    if not j or j == '000000' or j.startswith('#'):
        return None, None
    if INS.search(j) or UTIL.search(j):
        return None, None                       # 보험료·공과금·세금 = 명의 아님
    if ORG.search(j):
        return None, BANKPFX.sub('', j).strip()  # 조직 → belongs_to (은행접두 제거: 농협피어스미디어→피어스미디어)
    core = BANKPFX.sub('', j).strip()           # 은행접두 제거(기업김철중→김철중)
    if BIZ.search(core):
        return None, None                       # 상호·업종·메모(라온베베·와이헤어·월세) = 명의 아님
    if re.fullmatch(r'[가-힣]{2,4}', core):
        return core, None                       # 순수 사람이름만
    return None, None                           # 숫자/영문/메모/업종 → 공란


def owner_of(title, sheetname):
    t = str(title or '')
    m = re.search(r'([가-힣]{2,4})\s*명의', t)
    name = m.group(1) if m else None
    if not name:
        m2 = re.search(r'([가-힣]{2,4})\s*$', N(sheetname))
        name = m2.group(1) if m2 else None
    am = re.search(r"(\d[\d\-]{8,})", t)
    acct = norm_num(am.group(1)) if am else None
    return name, acct


def map_hdr(rows):
    hi = max(range(min(6, len(rows))), key=lambda i: sum(
        1 for c in rows[i] if c and any(k in str(c) for k in
        ['거래일', '입금', '지급', '출금', '적요', '상대계좌', '거래금액', '거래구분', '구분', '상태'])))
    h = [str(c).strip() if c else '' for c in rows[hi]]

    def f(*kw):
        for i, x in enumerate(h):
            if any(k in x for k in kw):
                return i
        return None
    return {
        'date': f('거래일자', '거래일'),
        'in': f('입금금액'), 'out': f('지급금액', '출금금액'),
        'gubun': f('거래구분', '구분', '상태'), 'amt1': f('거래금액'),
        'jeokyo': f('적요', '의뢰인'),
        'peer_acct': f('상대계좌번호', '이체상대계좌', '상대계좌'),
    }, hi


def amounts(cm, cell):
    if cm['in'] is not None and cm['out'] is not None:
        return to_int(cell('in')), to_int(cell('out'))
    g = str(cell('gubun') or '')
    amt = to_int(cell('amt1')) if cm['amt1'] is not None else to_int(cell('in'))
    if '입' in g:
        return amt, 0
    if '출' in g or '지급' in g:
        return 0, amt
    if cm['in'] is not None:
        return to_int(cell('in')), 0
    return 0, 0


def collect(d5):
    picked = {}
    for r, _, fs in os.walk(d5):
        if '030' not in N(r):
            continue
        for f in fs:
            fn = N(f)
            if not fn.endswith('.xlsx') or f.startswith(('.', '~')):
                continue
            for w in WANT:
                if w in fn:
                    prio = 2 if '_v3' in fn else (1 if '_수정' in fn else 0)
                    if w not in picked or prio > picked[w][0]:
                        picked[w] = (prio, os.path.join(r, f))
    return [v[1] for v in picked.values()]


def build(paths):
    import openpyxl
    nodes, edges = {}, []
    stats = defaultdict(int)

    def node(label, key, props):
        k = (label, key); nodes.setdefault(k, {})
        nodes[k].update({a: b for a, b in props.items() if b != ''}); return k

    for p in paths:
        wb = openpyxl.load_workbook(p, data_only=True)
        for sn in wb.sheetnames:
            ws = wb[sn]
            if ws.sheet_state != 'visible':
                continue
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 3:
                continue
            cm, hi = map_hdr(rows)
            if cm['peer_acct'] is None:
                continue
            base_owner, base_acct = owner_of(rows[0][0] if rows[0] else '', sn)
            cur_owner, cur_acct = base_owner, base_acct
            agg = defaultdict(lambda: {'n': 0, 'amt': 0, 'first': '', 'last': '', 'pnm': '', 'porg': ''})

            for r in rows[hi + 1:]:
                def cell(k):
                    i = cm.get(k)
                    return r[i] if i is not None and i < len(r) else None
                pa = norm_num(cell('peer_acct'))
                # 명의 구분행(국민 특이사항: 시트 중간 "OOO 명의 …") — 상대계좌 없는 텍스트행
                if not (pa and pa.isdigit() and len(pa) >= 8):
                    jt = ' '.join(str(c) for c in r if c)
                    mo = re.search(r'([가-힣]{2,4})\s*명의', jt)
                    if mo:
                        cur_owner = mo.group(1)
                        am = re.search(r"(\d[\d\-]{8,})", jt)
                        cur_acct = norm_num(am.group(1)) if am else None
                    continue
                din, dout = amounts(cm, cell)
                if din > 0:
                    direction, amt = 'in', din
                elif dout > 0:
                    direction, amt = 'out', dout
                else:
                    continue
                pnm, porg = clean_peer(str(cell('jeokyo') or ''))
                dt = str(cell('date')).strip()[:10] if cell('date') else ''
                a = agg[(cur_owner or '?', cur_acct or '', pa, direction)]
                a['n'] += 1; a['amt'] += amt
                a['pnm'] = pnm or a['pnm']; a['porg'] = porg or a['porg']
                a['first'] = min(a['first'] or dt, dt) if dt else a['first']
                a['last'] = max(a['last'], dt) if dt else a['last']

            for (own, oacct, pa, direction), a in agg.items():
                if own == '?':
                    stats['no_owner'] += 1
                    continue
                okey = oacct or f'명의:{own}'
                ok = node('vt_bacnt', okey, {'account_no': okey, 'dpstr': own, 'source_id': SRC})
                # 본계좌 명의 → 인물 노드 + has_account (EP1/EP2 정합·인물중심 노드셋)
                own_psn = node('vt_psn', own, {'name': own, 'source_id': SRC})
                edges.append(('has_account', own_psn, ok, {'source_id': SRC}))
                pk = node('vt_bacnt', pa, {'account_no': pa, 'dpstr': a['pnm'], 'source_id': SRC})
                if a['pnm']:
                    peer_psn = node('vt_psn', a['pnm'], {'name': a['pnm'], 'source_id': SRC})
                    edges.append(('has_account', peer_psn, pk, {'source_id': SRC}))
                props = {'txn_count': a['n'], 'total_amount': a['amt'],
                         'first_dlng_dt': a['first'], 'last_dlng_dt': a['last'], 'source_id': SRC}
                if direction == 'in':
                    edges.append(('transferred_to', pk, ok, props))
                else:
                    edges.append(('transferred_to', ok, pk, props))
                if a['pnm']:
                    stats['peer_named'] += 1
                if a['porg']:
                    org = node('vt_org', a['porg'], {'org_name': a['porg'], 'source_id': SRC})
                    edges.append(('belongs_to', pk, org, {'source_id': SRC}))
                    stats['peer_org'] += 1
    return nodes, edges, stats


KP = {'vt_bacnt': 'account_no', 'vt_org': 'org_name', 'vt_psn': 'name'}


def merge_cypher(nodes, edges):
    stmts = []
    for (label, key), props in nodes.items():
        setp = ', '.join(f"n.{a} = '{esc(b)}'" for a, b in props.items() if b != '')
        stmts.append(f"MERGE (n:{label} {{{KP[label]}:'{esc(props[KP[label]])}'}})" + (f" SET {setp}" if setp else ""))
    for rel, (fl, fk), (tl, tk), props in edges:
        a = f"(a:{fl} {{{KP[fl]}:'{esc(nodes[(fl,fk)][KP[fl]])}'}})"
        b = f"(b:{tl} {{{KP[tl]}:'{esc(nodes[(tl,tk)][KP[tl]])}'}})"
        sp = ', '.join(f"e.{k2} = '{esc(v2)}'" for k2, v2 in props.items() if v2 != '')
        stmts.append(f"MATCH {a}, {b} MERGE (a)-[e:{rel}]->(b)" + (f" SET {sp}" if sp else ""))
    return stmts


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--graph', default=GRAPH)
    ap.add_argument('--dry-run', action='store_true')
    args = ap.parse_args()
    d5 = [os.path.join(BASE, x) for x in os.listdir(BASE) if N(x).startswith('EP5.')][0]
    paths = collect(d5)
    print('[대상 파일]')
    for p in paths:
        print('  -', N(os.path.basename(p)))
    nodes, edges, stats = build(paths)
    from collections import Counter
    print(f"[빌드] 노드 {len(nodes)} / 엣지 {len(edges)} · 통계 {dict(stats)}")
    print("  노드:", dict(Counter(l for (l, _) in nodes)))
    print("  엣지:", dict(Counter(r for (r, *_) in edges)))
    stmts = merge_cypher(nodes, edges)
    if args.dry_run:
        named = [(k[1], p.get('dpstr', '')) for k, p in nodes.items() if k[0] == 'vt_bacnt' and p.get('dpstr')]
        print("  본계좌 명의 샘플:", [d for _, d in named if not d[0].isdigit()][:20])
        return
    app = create_app()
    with app.app_context():
        conn = psycopg2.connect(**app.config['DB_CONFIG']); conn.autocommit = False
        cur = conn.cursor()
        try:
            safe_set_graph_path(cur, args.graph)
            for vl in sorted({l for (l, _) in nodes}):
                cur.execute(f"CREATE VLABEL IF NOT EXISTS {vl};")
            for el in sorted({r for (r, *_) in edges}):
                cur.execute(f"CREATE ELABEL IF NOT EXISTS {el};")
            for s in stmts:
                cur.execute(s)
            conn.commit()
            print(f"[적재 완료] {len(stmts)}개 MERGE → {args.graph}")
        except Exception as e:
            conn.rollback(); print(f"[롤백] {e}"); raise
        finally:
            conn.close()


if __name__ == '__main__':
    main()
