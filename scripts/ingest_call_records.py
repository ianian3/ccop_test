#!/usr/bin/env python3
"""휴대폰 통화내역(통신영장 회신 csv/xlsx) → 통화망 그래프 적재 (결정론, LLM 무관).

EP3 013 통화회신: 발신번호·착신번호·통화시각·통화초·접속IP → contacted 집계 + used_ip.
온톨로지 매핑 (V4.7):
  발신/착신 번호  → vt_telno
  통화           → (발신)-[contacted {channel:'call', call_count, total_dur_sec, first/last_dt}]->(착신)  집계
  접속 IP        → vt_ip ; (발신)-[used_ip]->(IP)
CSV(euc-kr)·xlsx 헤더 자동 매핑. 멱등(MERGE).
실행: python3 scripts/ingest_call_records.py --root <013휴대전화번호폴더> --graph ep3_graph [--dry-run]
"""
import argparse
import csv
import glob
import os
import re
import sys
from collections import defaultdict

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from app import create_app
from app.database import safe_set_graph_path
import psycopg2

GRAPH = 'ep3_graph'
SRC_ID = 'EP3-013-call'
PHONE_RE = re.compile(r'^(01[016789]\d{6,8}|070\d{7,8}|02\d{7,8}|0\d{8,10})$')
IPV4_RE = re.compile(r'^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$')


def esc(v):
    return str(v).replace("\\", "\\\\").replace("'", "''")


def norm_num(v):
    return re.sub(r"[-\s'\xa0]", '', str(v)) if v else ''


def read_rows(path):
    """csv(euc-kr/utf-8)·xlsx → 행 리스트(값)."""
    ext = path.lower().rsplit('.', 1)[-1]
    if ext == 'csv':
        for enc in ('utf-8-sig', 'euc-kr', 'cp949'):
            try:
                with open(path, encoding=enc) as f:
                    return [tuple(r) for r in csv.reader(f)]
            except (UnicodeDecodeError, LookupError):
                continue
        return []
    if ext == 'xlsx':
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        return list(ws.iter_rows(values_only=True))
    if ext == 'xls':
        try:
            import xlrd
            ws = xlrd.open_workbook(path).sheet_by_index(0)
            return [tuple(ws.cell_value(i, j) for j in range(ws.ncols)) for i in range(ws.nrows)]
        except Exception:
            return []   # 암호화 등 → 스킵
    return []


def map_cols(rows):
    hi = max(range(min(10, len(rows))), key=lambda i: sum(
        1 for c in rows[i] if c and any(k in str(c) for k in ['발신번호', '착신번호', '통화', 'IP'])))
    h = [str(c).strip() if c else '' for c in rows[hi]]
    def find(*kws):
        for i, x in enumerate(h):
            if any(k in x for k in kws):
                return i
        return None
    return {
        'src': find('발신번호'), 'dst': find('착신번호'),
        'dur': find('통화초', '사용시간', '통화시간'),
        'date': find('통화월일', '통화시작', '통화일'),
        'ip': find('IP번호', 'IP주소', 'IP'),
    }, hi


def build(root):
    import unicodedata
    allf = [p for p in glob.glob(root + '/*.csv') + glob.glob(root + '/*.xlsx') + glob.glob(root + '/*.xls')
            if not os.path.basename(p).startswith(('._', '~'))]
    # v3 재비식별 폴더: _v3 파일이 있으면 v3만 채택 + v3로 대체되지 않는 별도 통화내역(E90 등)만 유지.
    # 구 역발신/tongwha/sk브로드밴드(값이 v3로 재매핑됨)는 배제해 구값·신값 혼재 방지.
    has_v3 = any('_v3' in unicodedata.normalize('NFC', os.path.basename(p)) for p in allf)
    if has_v3:
        def keep(p):
            b = unicodedata.normalize('NFC', os.path.basename(p))
            if '_v3' in b:
                return True
            return not any(k in b for k in ['역발신', 'tongwha', '브로드밴드'])
        files = [p for p in allf if keep(p)]
    else:
        files = allf
    nodes = {}
    call_agg = defaultdict(lambda: {'n': 0, 'dur': 0.0, 'first': '', 'last': ''})
    ip_use = set()
    n_files, n_rec, n_skip = 0, 0, 0

    def node(label, key, props):
        k = (label, key); nodes.setdefault(k, {})
        nodes[k].update({a: b for a, b in props.items() if b}); return k

    for f in files:
        rows = read_rows(f)
        if not rows:
            n_skip += 1
            continue
        cm, hi = map_cols(rows)
        if cm['src'] is None or cm['dst'] is None:
            n_skip += 1
            continue
        n_files += 1
        def cell(r, k):
            i = cm.get(k)
            return r[i] if i is not None and i < len(r) else None
        for r in rows[hi + 1:]:
            s = norm_num(cell(r, 'src')); d = norm_num(cell(r, 'dst'))
            if not (PHONE_RE.match(s) and PHONE_RE.match(d)) or s == d:
                continue
            n_rec += 1
            dt = str(cell(r, 'date')).strip().strip("'")[:8] if cell(r, 'date') else ''
            try:
                dur = float(re.sub(r'[^\d.]', '', str(cell(r, 'dur') or 0)) or 0)
            except ValueError:
                dur = 0.0
            a = call_agg[(s, d)]
            a['n'] += 1; a['dur'] += dur
            a['first'] = min(a['first'] or dt, dt) if dt else a['first']
            a['last'] = max(a['last'], dt) if dt else a['last']
            ip = norm_num(cell(r, 'ip')) if cm.get('ip') is not None else ''
            ipv = str(cell(r, 'ip')).strip() if cm.get('ip') is not None and cell(r, 'ip') else ''
            if IPV4_RE.match(ipv):
                ip_use.add((s, ipv))

    edges = []
    for (s, d), a in call_agg.items():
        sk = node('vt_telno', s, {'telno': s, 'source_id': SRC_ID})
        dk = node('vt_telno', d, {'telno': d, 'source_id': SRC_ID})
        edges.append(('contacted', sk, dk, {'channel': 'call', 'call_count': a['n'],
                      'total_dur_sec': round(a['dur']), 'first_dt': a['first'],
                      'last_dt': a['last'], 'source_id': SRC_ID}))
    for (s, ip) in ip_use:
        sk = node('vt_telno', s, {'telno': s, 'source_id': SRC_ID})
        ik = node('vt_ip', ip, {'ip_addr': ip, 'source_id': SRC_ID})
        edges.append(('used_ip', sk, ik, {'source_id': SRC_ID}))
    return nodes, edges, n_files, n_rec, n_skip


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--root', required=True)
    ap.add_argument('--graph', default=GRAPH)
    ap.add_argument('--dry-run', action='store_true')
    args = ap.parse_args()
    nodes, edges, nf, nr, nsk = build(args.root)
    from collections import Counter
    print(f"[빌드] 통화파일 {nf}개(미지원/스킵 {nsk}) · 통화레코드 {nr:,} → 노드 {len(nodes)} / 엣지 {len(edges)}")
    print("  노드:", dict(Counter(l for (l, _) in nodes)))
    print("  엣지:", dict(Counter(r for (r, *_) in edges)))
    KP = {'vt_telno': 'telno', 'vt_ip': 'ip_addr'}
    stmts = []
    for (label, key), props in nodes.items():
        setp = ', '.join(f"n.{a} = '{esc(b)}'" for a, b in props.items())
        stmts.append(f"MERGE (n:{label} {{{KP[label]}:'{esc(props[KP[label]])}'}})" + (f" SET {setp}" if setp else ""))
    for rel, (fl, fk), (tl, tk), props in edges:
        a = f"(a:{fl} {{{KP[fl]}:'{esc(nodes[(fl,fk)][KP[fl]])}'}})"
        b = f"(b:{tl} {{{KP[tl]}:'{esc(nodes[(tl,tk)][KP[tl]])}'}})"
        sp = ', '.join(f"e.{k2} = '{esc(v2)}'" for k2, v2 in props.items() if v2 != '')
        stmts.append(f"MATCH {a}, {b} MERGE (a)-[e:{rel}]->(b)" + (f" SET {sp}" if sp else ""))
    if args.dry_run:
        for s in stmts[:5]:
            print("  ", s[:140])
        return
    app = create_app()
    with app.app_context():
        conn = psycopg2.connect(**app.config['DB_CONFIG']); conn.autocommit = False
        cur = conn.cursor()
        try:
            safe_set_graph_path(cur, args.graph)
            for vl in sorted({l for (l, _) in nodes}):
                cur.execute(f"CREATE VLABEL IF NOT EXISTS {vl};")
            for el in sorted({r for (r, *_) in edges}):
                cur.execute(f"CREATE ELABEL IF NOT EXISTS {el};")
            for s in stmts:
                cur.execute(s)
            conn.commit()
            print(f"[적재 완료] {len(stmts)}개 MERGE → {args.graph}")
        except Exception as e:
            conn.rollback(); print(f"[롤백] {e}"); raise
        finally:
            conn.close()


if __name__ == '__main__':
    main()
