#!/usr/bin/env python3
"""계좌거래내역(금융영장 회신 xlsx) → 자금흐름 그래프 적재 (결정론, LLM 무관).

EP3 012 금융회신: 본계좌의 거래내역(입금/지급·상대계좌·상대명) → transferred_to 집계.
온톨로지 매핑 (V4.7):
  본계좌            → vt_bacnt {account_no, bank_nm, dpstr}
  상대계좌          → vt_bacnt {account_no, dpstr}
  입금(상대→본)     → (상대)-[transferred_to {txn_count, total_amount, first/last_dlng_dt}]->(본)
  지급(본→상대)     → (본)-[transferred_to ...]->(상대)
  상대명이 조직(피어스미디어 등) → vt_org + (상대계좌)-[belongs_to]->(조직)  (선택)
상대계좌별로 방향·금액 집계(온톨로지 transferred_to = 요약 엣지). 멱등(MERGE).
실행: python3 scripts/ingest_account_txn.py --xlsx <f> --sheet 정리 --owner-name 김은희 --owner-bank 농협 --graph ep3_graph
"""
import argparse
import re
import sys
import os
from collections import defaultdict

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from app import create_app
from app.database import safe_set_graph_path
import psycopg2

GRAPH = 'ep3_graph'
ORG_HINT = re.compile(r'(미디어|컴퍼니|㈜|주식회사|인터내셔널|트레이드|커머스|엔터|글로벌|테크|시스템)')


def esc(v):
    return str(v).replace("\\", "\\\\").replace("'", "''")


def norm_num(v):
    return re.sub(r'[-\s\xa0]', '', str(v)) if v else ''


def to_int(v):
    n = norm_num(v)
    return int(n) if n.isdigit() else 0


def _map(rows):
    hi = max(range(min(5, len(rows))), key=lambda i: sum(1 for c in rows[i] if c and any(
        k in str(c) for k in ['거래일', '입금', '지급', '계좌번호', '입출금명', '거래시간'])))
    h = [str(c).strip() if c else '' for c in rows[hi]]
    def find(*kws):
        for i, x in enumerate(h):
            if any(k in x for k in kws):
                return i
        return None
    return {
        'date': find('거래일자', '거래일시', '거래년월일', '거래일'),
        'time': find('거래시간', '거래일시분초'),
        'in': find('입금금액', '입금'),
        'out': find('지급금액', '출금금액'),
        'gubun': find('거래구분', '구분', '입출금'),   # 입금/지급 방향 코드
        'amt1': find('거래금액'),                       # 단일 거래금액(방향은 gubun)
        'peer_nm': find('입출금명', '상대명', '적요', '취급'),
        'peer_acct': find('상대계좌번호', '계좌번호'),   # 상대계좌 우선
        'bank': find('상대은행', '송금은행', '거래점', '은행'),
    }, hi


def build(xlsx, sheet, owner_name, owner_bank, src_id='EP3-012'):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb[sheet] if sheet else next((wb[x] for x in wb.sheetnames if wb[x].sheet_state=='visible'), wb.active)
    if ws.sheet_state != 'visible':
        raise SystemExit(f"'{sheet}'는 숨김 시트(state={ws.sheet_state}).")
    rows = list(ws.iter_rows(values_only=True))
    cm, hi = _map(rows)
    has_amt = (cm['in'] is not None and cm['out'] is not None) or \
              (cm['gubun'] is not None and (cm['amt1'] is not None or cm['in'] is not None))
    if cm['peer_acct'] is None or not has_amt:
        raise SystemExit(f"필수 컬럼 미발견: {cm}")

    def amounts(cell_fn):
        """양식별 (입금액, 지급액) 판정. ① 입금/지급 분리 ② 구분+단일금액 ③ 구분+입금금액."""
        if cm['in'] is not None and cm['out'] is not None:
            return to_int(cell_fn('in')), to_int(cell_fn('out'))
        g = str(cell_fn('gubun') or '')
        amt = to_int(cell_fn('amt1')) if cm['amt1'] is not None else to_int(cell_fn('in'))
        if any(x in g for x in ['입금', '입']):
            return amt, 0
        if any(x in g for x in ['지급', '출금', '출']):
            return 0, amt
        # 거래구분이 방향이 아님(채널 '인터넷' 등) + 입금금액 컬럼만 존재 → 입금 전용 내역
        if cm['in'] is not None and cm['out'] is None:
            return to_int(cell_fn('in')), 0
        return 0, 0

    owner_key = f'{owner_bank}-{owner_name}'          # 본계좌 번호 미상 → 명의 기반 임시키
    nodes = {('vt_bacnt', owner_key): {'account_no': owner_key, 'bank_nm': owner_bank,
                                       'dpstr': owner_name, 'source_id': src_id},
             ('vt_psn', owner_name): {'name': owner_name, 'source_id': src_id}}  # 명의 → 인물 노드
    # (상대계좌, 방향) → 집계
    agg = defaultdict(lambda: {'n': 0, 'amt': 0, 'first': '', 'last': '', 'nm': ''})
    orgs = {}

    def cell(r, k):
        i = cm.get(k)
        return r[i] if i is not None and i < len(r) else None

    for r in rows[hi + 1:]:
        pa = norm_num(cell(r, 'peer_acct'))
        if not pa or not pa.isdigit() or len(pa) < 8:
            continue
        nm = str(cell(r, 'peer_nm')).strip() if cell(r, 'peer_nm') else ''
        din, dout = amounts(lambda k: cell(r, k))
        dt = str(cell(r, 'date')).strip()[:10] if cell(r, 'date') else ''
        if din > 0:
            direction, amt = 'in', din       # 상대 → 본
        elif dout > 0:
            direction, amt = 'out', dout      # 본 → 상대
        else:
            continue
        k = (pa, direction)
        a = agg[k]
        a['n'] += 1; a['amt'] += amt; a['nm'] = nm or a['nm']
        a['first'] = min(a['first'] or dt, dt) if dt else a['first']
        a['last'] = max(a['last'], dt) if dt else a['last']

    # 본계좌 명의 → has_account (인물중심 노드셋 · EP1/EP2 정합)
    edges = [('has_account', ('vt_psn', owner_name), ('vt_bacnt', owner_key), {'source_id': src_id})]
    for (pa, direction), a in agg.items():
        peer = ('vt_bacnt', pa)
        nodes.setdefault(peer, {})
        nodes[peer].update({'account_no': pa, 'dpstr': a['nm'], 'source_id': src_id})
        props = {'txn_count': a['n'], 'total_amount': a['amt'],
                 'first_dlng_dt': a['first'], 'last_dlng_dt': a['last'], 'source_id': src_id}
        if direction == 'in':
            edges.append(('transferred_to', peer, ('vt_bacnt', owner_key), props))
        else:
            edges.append(('transferred_to', ('vt_bacnt', owner_key), peer, props))
        # 상대명이 조직이면 belongs_to, 사람이름이면 인물+has_account
        if a['nm'] and ORG_HINT.search(a['nm']):
            ok = ('vt_org', a['nm'])
            nodes.setdefault(ok, {'org_name': a['nm'], 'source_id': src_id})
            edges.append(('belongs_to', peer, ok, {'source_id': src_id}))
        elif a['nm'] and re.fullmatch(r'[가-힣]{2,4}', a['nm'].strip()):
            pp = ('vt_psn', a['nm'].strip())
            nodes.setdefault(pp, {'name': a['nm'].strip(), 'source_id': src_id})
            edges.append(('has_account', pp, peer, {'source_id': src_id}))
    return nodes, edges


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--xlsx', required=True)
    ap.add_argument('--sheet', default=None)
    ap.add_argument('--owner-name', required=True)
    ap.add_argument('--owner-bank', required=True)
    ap.add_argument('--graph', default=GRAPH)
    ap.add_argument('--src-id', default='EP3-012')
    ap.add_argument('--dry-run', action='store_true')
    args = ap.parse_args()

    nodes, edges = build(args.xlsx, args.sheet, args.owner_name, args.owner_bank, args.src_id)
    from collections import Counter
    print(f"[빌드] 노드 {len(nodes)} / 엣지 {len(edges)}")
    print("  노드:", dict(Counter(l for (l, _) in nodes)))
    print("  엣지:", dict(Counter(r for (r, *_) in edges)))
    KP = {'vt_bacnt': 'account_no', 'vt_org': 'org_name', 'vt_psn': 'name'}
    stmts = []
    for (label, key), props in nodes.items():
        setp = ', '.join(f"n.{a} = '{esc(b)}'" for a, b in props.items() if b != '')
        stmts.append(f"MERGE (n:{label} {{{KP[label]}:'{esc(props[KP[label]])}'}})" + (f" SET {setp}" if setp else ""))
    for rel, (fl, fk), (tl, tk), props in edges:
        a = f"(a:{fl} {{{KP[fl]}:'{esc(nodes[(fl,fk)][KP[fl]])}'}})"
        b = f"(b:{tl} {{{KP[tl]}:'{esc(nodes[(tl,tk)][KP[tl]])}'}})"
        sp = ', '.join(f"e.{k2} = '{esc(v2)}'" for k2, v2 in props.items() if v2 != '')
        stmts.append(f"MATCH {a}, {b} MERGE (a)-[e:{rel}]->(b)" + (f" SET {sp}" if sp else ""))
    if args.dry_run:
        for s in stmts[:5]:
            print("  ", s[:150])
        return
    app = create_app()
    with app.app_context():
        conn = psycopg2.connect(**app.config['DB_CONFIG']); conn.autocommit = False
        cur = conn.cursor()
        try:
            safe_set_graph_path(cur, args.graph)
            for vl in sorted({l for (l, _) in nodes}):
                cur.execute(f"CREATE VLABEL IF NOT EXISTS {vl};")
            for el in sorted({r for (r, *_) in edges}):
                cur.execute(f"CREATE ELABEL IF NOT EXISTS {el};")
            for s in stmts:
                cur.execute(s)
            conn.commit()
            print(f"[적재 완료] {len(stmts)}개 MERGE → {args.graph}")
        except Exception as e:
            conn.rollback(); print(f"[롤백] {e}"); raise
        finally:
            conn.close()


if __name__ == '__main__':
    main()
